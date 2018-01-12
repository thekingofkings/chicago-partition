#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Dec 11 21:02:07 2017

@author: kok

Generate tract-level features for regressions
"""


from openpyxl import load_workbook
import pandas as pd
import scipy.stats as stats
import numpy as np


def retrieve_corina_features():
    from pandas import read_stata
    
    r = read_stata('data/SE2000_AG20140401_MSAcmsaID.dta')
    cnt = 0
    header = ['pop00', 'ppov00', 'disadv00', 'pdensmi00', 'hetero00', 'phisp00', 'pnhblk00']
    
    fields_dsp = ['total population', 'poverty index', 'disadvantage index', 
                  'population density', 'ethnic diversity', 'pct hispanic', 'pct black']

    ST = {}
    for row in r.iterrows():
        tract = row[1]
        if tract['statetrim'] == '17' and tract['countrim'] == '031':
            cnt += 1
            tl = []
            tid = '17031' + tract['tracttrim']
            for h in header:
                tl.append(tract[h])
            ST[tid] = tl
    return fields_dsp, ST


def retrieve_income_features():
    """
    read the xls file '../data/Household Income by Race and Census Tract and Community Area.xlsx'
    """
    wb = load_workbook('data/Household Income by Race and Census Tract and Community Area.xlsx')
    ws = wb.active
    tractColumn = [cell.value for cell in ws['h']]
    dataColumns = ws['K1:DU890']
    
    header = []
    header_description = []
    income_features = []
    tractIDs = []
    for idx, tractID in enumerate(tractColumn):
        if idx == 0:
            header = [cell.value for cell in dataColumns[idx]]
        elif idx == 1:
            header_description = [cell.value for cell in dataColumns[idx]]
        elif idx == 2:
            header_description = ["{} {}".format(header_description[i], cell.value)
                                  for i, cell in enumerate(dataColumns[idx])]
        else:
            if tractID != None:
                tractIDs.append(int("17031"+tractID))
                row = [cell.value for cell in dataColumns[idx]]
                income_features.append(row)
    featureDF = pd.DataFrame(data=income_features, index=tractIDs, columns=header)
    header_decode = dict(zip(header, header_description))
    return featureDF, header_decode

def retrieve_income_features_entropy(featureDF):
    '''compute features including diversity and percentage'''
    featureDF = featureDF.loc[:,~featureDF.columns.duplicated()]
    added_features = []
    # calculate percentage for different ethnic population
    ethnics = ['H','B','I','D']

    ##calculate overall population
    featureDF['B1901Z01'] = 0
    for ethnic in ethnics:
        feature_col_name = 'B1901%s01'%ethnic
        featureDF['B1901Z01'] += featureDF['%s'%feature_col_name]
    ##calcluate percentage
    for ethnic in ethnics:
        feature_col_name = 'B1901%s01'%ethnic
        featureDF['%s_pct'%feature_col_name] = featureDF.apply(lambda row: _percentage(row,feature_col_name,'B1901Z01'),axis=1)
    ##calcaulate population diversity
    entropy_features_pop = ['B1901%s01_pct'%ethnic for ethnic in ethnics]
    featureDF['pop_diversity'] = featureDF.apply(lambda row: _entropy(row,entropy_features_pop), axis=1)

    # calculate everage income of different ethnics
    index_to_income = [5000,12500,17500,22500,27500,32500,37500,42500,47500,55000,67500,87500,112500,137500,175000,250000]

    for ethnic in ethnics:
        weight_features = ['B1901%s%s'%(ethnic, str(level).zfill(2)) for level in range(2,18)]
        featureDF['mean_%s01'%ethnic] = featureDF.apply(lambda row: _weighted_mean(row,index_to_income,weight_features), axis=1)

    entropy_features_income = ['mean_%s01'%ethnic for ethnic in ethnics]
    featureDF['income_diversity'] = featureDF.apply(lambda row: _entropy(row, entropy_features_income), axis=1)


    added_features.append('B1901Z01')
    for ethnic in ethnics:
        feature_col_name = 'B1901%s01' % ethnic
        added_features.append('%s_pct'%feature_col_name)
        added_features.append('mean_%s01'%ethnic)
    added_features.append('pop_diversity')
    added_features.append('income_diversity')
    return featureDF,added_features

def _percentage(row,numerator_feature,denominator):
    return 0 if np.isnan(row[numerator_feature]/float(row[denominator]))  else row[numerator_feature]/float(row[denominator])


def _entropy(row,features):
    x = row[features].values
    return stats.entropy(x) if stats.entropy(x) != float('-inf') else 0

def _weighted_mean(row, values, weight_features):
    weights = row[weight_features].values
    if weights.sum()!= 0:
        return np.average(values,weights=weights)
    else:
        return 0

def retrieve_income_features_entropy_community(featureDF, tract_to_community_mapping):
    """
    now you may not need to use this,compute for all tracts from different communities
    :param featureDF:
    :param tract_to_community_mapping:
    :return:
    """
    featureDF = pd.merge(featureDF,tract_to_community_mapping,how='inner',left_index=True,right_index=True)
    groupDF = featureDF.groupby(by='CA_no').sum()
    groupDF = groupDF.reset_index().set_index('CA_no')
    groupDF = retrieve_income_features_entropy(groupDF)
    featureDF_with_groupinfo = pd.merge(featureDF,groupDF,how='left',left_on='CA_no',right_index=True,suffixes=['','_CA'])
    return featureDF_with_groupinfo

def retrieve_crime_count(year=2010):
    """
    Output:
        Dataframe of crime counts in various categoires. The index type is integer.
        The column name is crime type
    """
    Y = pd.read_csv("data/chicago-crime-tract-level-{0}.csv".format(year),
                    header=0, index_col=0)
    return Y
        


def validate_region_keys():
    """
    We have three sources of tract IDs: 
        1) shapefile probably from 2010
        2) the census demographics from 2000 (Corina provides)
        3) the census demographics from 2010 (download from web)
    
    The question to figure out is which year the shapefile is from? - confirmed from 2010.
    Further, are there any missing tracts in the shapefile? - yes. there are 8
    """
    from tract import Tract
    tracts = Tract.createAllTracts()
    shp_keys = set(tracts.keys())
    
    fields_dsp, st = retrieve_corina_features() # 2000 census data
    corina_keys = set(st.keys())
    
    f, d = retrieve_income_features()  # 2010 census data
    census2010_keys = set(f.index)
    
    print "Compare the tract IDs between shapefile and 2000 census"
    print "len(shp) {}, len(2000 census) {}".format(len(shp_keys), len(corina_keys))
    print "intersection: {}, union {}.".format(len(shp_keys & corina_keys), 
                         len(shp_keys | corina_keys))
    
    print "Compare the tract IDs between shapefile and 2010 census"
    print "len(shp) {}, len(2010 census) {}".format(len(shp_keys), len(census2010_keys))
    print "intersection: {}, union {}.".format(len(shp_keys & census2010_keys), 
                         len(shp_keys | census2010_keys))

    tract_to_community_mapping = {'CA_no':[],'tract_no':[]}
    for s in census2010_keys:
        if s not in shp_keys:
            print s

if __name__ == '__main__':
    validate_region_keys()
    f,d = retrieve_income_features()
    f,added_features = retrieve_income_features_entropy(f)
    y = retrieve_crime_count()